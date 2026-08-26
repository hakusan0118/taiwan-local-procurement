#!/usr/bin/env python3
"""Build private Excel analysis workbooks from public procurement CSV files."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


CASE_NUMBER_FIELDS = {"year", "budget_amount", "award_amount", "winner_count"}
VENDOR_NUMBER_FIELDS = {"year", "case_award_amount", "winner_count"}
MONEY_FIELDS = {
    "budget_amount", "award_amount", "case_award_amount", "total_award_amount",
    "average_award_amount", "total_case_award_amount",
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def read_csv(path: Path, number_fields: set[str]) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    for row in rows:
        for field in number_fields:
            value = row.get(field, "")
            if value not in (None, ""):
                try:
                    row[field] = int(float(value))
                except ValueError:
                    pass
    return rows


def parse_years(values: list[str]) -> list[int]:
    """接受 2023、2010 2011、2010~2023、2010-2023 等輸入。"""
    years: set[int] = set()
    tokens = re.split(r"[\s,，]+", " ".join(values).strip())
    for token in filter(None, tokens):
        match = re.fullmatch(r"(20\d{2})\s*[~～\-–—]\s*(20\d{2})", token)
        if match:
            start, end = map(int, match.groups())
            if start > end:
                raise ValueError(f"年度範圍起點不得晚於終點：{token}")
            years.update(range(start, end + 1))
        elif re.fullmatch(r"20\d{2}", token):
            years.add(int(token))
        else:
            raise ValueError(f"無法辨識年度：{token}；請輸入 2023 或 2010~2023")
    if not years:
        raise ValueError("至少需要一個年度")
    return sorted(years)


def add_sheet(workbook: Workbook, title: str, rows: list[dict], table_name: str) -> None:
    sheet = workbook.create_sheet(title)
    if not rows:
        sheet.append(["沒有資料"])
        return
    headers = list(rows[0])
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A2"
    table = Table(displayName=table_name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    sheet.add_table(table)
    for index, header in enumerate(headers, 1):
        values = [str(sheet.cell(row, index).value or "") for row in range(1, min(sheet.max_row, 300) + 1)]
        width = min(max(max(map(len, values)) + 2, 10), 42)
        if header in MONEY_FIELDS:
            # 金額保留為真正數值，使用一般數字＋千分位，不套貨幣符號。
            # 至少 18 字元寬，避免 Excel 因欄寬不足顯示 #####。
            width = max(width, 18)
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, index).number_format = '#,##0'
        sheet.column_dimensions[get_column_letter(index)].width = width


def agency_stats(cases: list[dict]) -> list[dict]:
    grouped: dict[str, dict[str, object]] = defaultdict(lambda: {"count": 0, "amounts": []})
    for row in cases:
        name = str(row.get("unit_name", ""))
        grouped[name]["count"] = int(grouped[name]["count"]) + 1
        amount = row.get("award_amount")
        if isinstance(amount, int):
            grouped[name]["amounts"].append(amount)
    result = [{
        "unit_name": name,
        "case_count": values["count"],
        "total_award_amount": sum(amounts),
        "average_award_amount": round(sum(amounts) / len(amounts)) if amounts else 0,
    } for name, values in grouped.items() for amounts in [values["amounts"]]]
    return sorted(result, key=lambda row: row["total_award_amount"], reverse=True)


def annual_stats(cases: list[dict]) -> list[dict]:
    grouped: dict[int, list[int]] = defaultdict(list)
    counts: dict[int, int] = defaultdict(int)
    for row in cases:
        year = int(row["year"])
        counts[year] += 1
        amount = row.get("award_amount")
        if isinstance(amount, int):
            grouped[year].append(amount)
    return [{
        "year": year,
        "case_count": counts[year],
        "total_award_amount": sum(grouped[year]),
        "average_award_amount": round(sum(grouped[year]) / len(grouped[year])) if grouped[year] else 0,
    } for year in sorted(counts)]


def vendor_stats(vendors: list[dict]) -> list[dict]:
    cases_by_vendor: dict[str, dict[str, int | None]] = defaultdict(dict)
    for row in vendors:
        name = str(row.get("vendor_name", ""))
        case_key = str(row.get("case_key", ""))
        amount = row.get("case_award_amount")
        if name and case_key:
            cases_by_vendor[name][case_key] = amount if isinstance(amount, int) else None
    result = [{
        "vendor_name": name,
        "case_count": len(cases),
        "total_case_award_amount": sum(amount for amount in cases.values() if amount is not None),
        "note": "多家共同得標案件的總決標金額不可視為單一廠商實得金額",
    } for name, cases in cases_by_vendor.items()]
    return sorted(result, key=lambda row: row["total_case_award_amount"], reverse=True)


REGION_LABELS = {
    "hualien": "花蓮縣",
    "taichung": "臺中市",
}


def build(years: list[int], data_root: Path, output_dir: Path, region: str = "hualien") -> Path:
    source = data_root / "processed" / region
    requested = sorted(set(years))
    cases: list[dict] = []
    available: list[int] = []
    for year in requested:
        annual = [
            row for row in read_csv(source / f"procurement_{year}.csv", CASE_NUMBER_FIELDS)
            if row.get("year") == year
        ]
        if annual:
            cases.extend(annual)
            available.append(year)
    missing = [year for year in requested if year not in available]
    if not cases:
        raise ValueError(f"找不到指定年度案件 CSV：{', '.join(map(str, requested))}")

    available_set = set(available)
    vendors = [
        row for row in read_csv(source / "vendors.csv", VENDOR_NUMBER_FIELDS)
        if row.get("year") in available_set
    ]
    quality = [
        row for row in read_csv(source / "data_quality.csv", {"year"})
        if row.get("year") in available_set
    ]
    scope = [{
        "requested_years": f"{requested[0]}~{requested[-1]}" if len(requested) > 1 else str(requested[0]),
        "included_years": "、".join(map(str, available)),
        "missing_years": "、".join(map(str, missing)) if missing else "無",
        "note": "缺少年度不會納入統計；補齊公開 CSV 後重新執行即可。",
    }]

    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "資料範圍", scope, "DataScope")
    add_sheet(workbook, "年度統計", annual_stats(cases), "AnnualStats")
    add_sheet(workbook, "案件主表", cases, "CasesAll")
    add_sheet(workbook, "得標廠商明細", vendors, "VendorsAll")
    add_sheet(workbook, "機關統計", agency_stats(cases), "AgenciesAll")
    add_sheet(workbook, "廠商統計", vendor_stats(vendors), "VendorStatsAll")
    add_sheet(workbook, "資料品質", quality, "QualityAll")
    output_dir.mkdir(parents=True, exist_ok=True)
    label = f"{requested[0]}-{requested[-1]}" if len(requested) > 1 else str(requested[0])
    output = output_dir / f"{REGION_LABELS[region]}_{label}_決標分析.xlsx"
    workbook.save(output)
    if missing:
        print(f"警告：缺少年度 {', '.join(map(str, missing))}")
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--years", nargs="+", required=True)
    parser.add_argument("--data-root", type=Path, default=Path("data"))
    parser.add_argument("--output-dir", type=Path, default=Path("excel-output"))
    parser.add_argument("--region", choices=sorted(REGION_LABELS), default="hualien")
    args = parser.parse_args()
    print(build(parse_years(args.years), args.data_root, args.output_dir, args.region))


if __name__ == "__main__":
    main()
